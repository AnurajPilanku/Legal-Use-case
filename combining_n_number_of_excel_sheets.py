#Excel sheet Modeling for Legal usecase
#Anuraj Pilanku
#importing modules
import openpyxl
import sys
input=sys.argv[1]
output=sys.argv[2]
win = openpyxl.load_workbook(input)

#geting names of all the sheets from an excel workbook and storing it in a list
all_sheetnames=win.sheetnames

#converting string to variable and assigning value to the variable
for i in range(0,len(all_sheetnames)):
    globals()["zhe"+str(i)]=win[all_sheetnames[i]]
#getting all the variables in a python code
variablelist=list()
all_variables=dir()#taking all the data in diectory(python script)
for name in all_variables:
    if name.startswith("z"):#taking the variables starting with "z" and appending it to a adictionory
        print(name,type(name))
        variablelist.append(name)

wout =openpyxl.Workbook()
sout = wout.active
mr=globals()[variablelist[0]].max_row
mc=globals()[variablelist[0]].max_column
if len(variablelist)>1:
    for j in range(1,len(variablelist)):
        for i in range(1,mc+1):
            sout.cell(row=j+2, column=i).value = globals()[variablelist[j]].cell(row=2, column=i).value
for i in range(1,3):
    for j in range(1,mc+1):#since only two columns are  needed
        sout.cell(row=i, column=j).value=globals()[variablelist[0]].cell(row=i, column=j).value
wout.save(output)

#python pandas.py C:\Users\2040664\anuraj\LEGAL\numberofsheets.xlsx C:\Users\2040664\anuraj\LEGAL\




