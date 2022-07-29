#Excel sheet Modeling for Legal usecase
#Anuraj Pilanku
#importing modules
import string
import openpyxl
import sys
from openpyxl.styles import Alignment, Border,Side,Font,PatternFill
from copy import copy
from colorama import *

#inputs and outputs
input=sys.argv[1]
input2=sys.argv[2]
input3=sys.argv[3]
path=sys.argv[4]

#Hardcoreded
laia_shift={"LAIA Eagle Import":"Night","LAIA Scheduled Notification":"Night","LAIA Sync BUs from EMU":"EB","LAIA-EID Data Sync":"EB","LAIA BU Export":"EB","LAIA BU IMPORT":"EB","LAIA Eagle Report Import":"EB","LAIA Eagle Import Backup":"EB","LAIA Reminder Monitoring":"Noon","LAIA Eagle Export":"Noon","LAIA - BU Export_AM":"Noon","LAIA - BU Import_AM":"Noon"}
laia_servertime={"LAIA Eagle Import":"Daily 12.30 PM (Job disabled/not configures in 289C server)","LAIA Scheduled Notification":"Daily 5.30 PM ","LAIA Sync BUs from EMU":"Daily 10.30 PM ","LAIA-EID Data Sync":"Daily 10.30 PM ","LAIA - BU Export":"Daily 10.30 PM ","LAIA - BU Import":"Daily 11:00 PM","LAIA Eagle Report Import":"Daily 11.15 PM ","LAIA Eagle Import Backup":"Daily 1.15 AM ","LAIA Reminder Monitoring":"Weekly - Monday - 6 AM ","LAIA Eagle Export":"Mon to Friday 9:00:00 AM (Job disabled/not configures in 289C server)","LAIA - BU Export_AM":"Daily 9.50 AM ","LAIA - BU Import_AM":"Daily 10.20 AM ","LAIA Sync BUs from EMU":"Daily 10.30 PM","LAIA Reminder Monitoring":"Weekly - Monday - 6 AM","LAIA - BU Import_AM":"Daily 10.20 AM","LAIA BU Export":"Daily 10.30 PM","LAIA BU IMPORT":"Daily 11:00 PM","LAIA BU IMPORT AM":"Daily 10.20 AM","LAIA_Sync_BUs_fromEMU":"Daily 10.30 PM","LAIA - CaseDetailsMailsend":" "}
laps_shift={"LAPS_EID_Integration_WD":"EB","LAPS-CustodianInfo_WD":"EB","LAPS-LegalHoldCustodian":"EB","LAPS_Populate_UPN_WD":"Noon","LAPS-LegalHoldCustodian_WD":"Noon","LAPS-Employee Status Change Report_WD":"Noon","LAPS-MailManagementLegalHoldPopulate":"Noon","LAPS-IPP Feed":"Noon","LAPS-LegalHoldCustodian_AccessRemoval_Notif":"Noon","LAPS-MailManage-ValidateHolds":"Noon"}
laps_servertime={"LAPS_EID_Integration_WD":"Monday to Friday 12:10 AM","LAPS-CustodianInfo_WD":"Monday to Friday 1:30 AM","LAPS-LegalHoldCustodian":"Daily 3:00 AM(Job disabled)","LAPS_Populate_UPN_WD":"Daily 3:30 AM","LAPS-LegalHoldCustodian_WD":"Daily 3:30 AM","LAPS-Employee Status Change Report_WD":"Every Wednesday 4:00 AM","LAPS-MailManagementLegalHoldPopulate":"Monday to Friday 5:00 AM","LAPS-IPP Feed":"Every Monday 7:00 AM","LAPS-LegalHoldCustodian_AccessRemoval_Notif":"Monday to Friday 8:00 AM","LAPS-MailManage-ValidateHolds":"Every 21 days 8:00 AM","LAPS-CustodianInfo_WD":"Monday to Friday 1:30 AM","LAPS-LegalHoldCustodian_WD":"Daily 3:30 AM","LAPS-MailManagementLegalHoldPopulate":"Monday to Friday 5:00 AM","LAPS-LegalHoldCustodian_AccessRemoval_Notif":"Monday to Friday 8:00 AM","Store Procedure [SP_MailManagementLegalHoldPopulate] to run everyday at 5 AM CST":"Everyday at 5 AM","LAPS - EID Integration":"Monday to Friday 12:10 AM"}
lacc_servertime={"LRN - BCCM_Navax_File_Transfer":"Weekly-Monday-7.30 AM","BCCM-Navex EID Integration":"Weekly-Monday-9 PM","LRN - Clean up LAW connections":"Everyday every 1 hr between 12 AM to 11.59 PM","LACC_3M_HR":"Daily 6.30 PM","ENTERPRISE_EMPLOYEE_UPDATE_EXTRACT":"Weekly-Monday-5.30 PM","ENTERPRISE_HR_IMPORT":"Weekly-Monday-4 PM"}
lacc_shift={}
#creating workbook and sheets
#workbook for laia
wb=openpyxl.load_workbook(input)
sh=wb.active
mc=sh.max_row
mr=sh.max_column
#workbook for laps
wp=openpyxl.load_workbook(input2)
sp=wp.active
mp=sp.max_row
ml=sp.max_column
#workbook for lacc
wc=openpyxl.load_workbook(input3)
sc=wc.active
mz=sc.max_row
mn=sc.max_column

w=openpyxl.Workbook()
lai=w.active
#renaming sheet
lai.title="LAIA"
#creating a new sheet in an existing workbook
#laps
lap=w.create_sheet("Sheet_B",1)
lap.title="LAPS"
#lacc
lacc=w.create_sheet("Sheet_C",2)
lacc.title="LACC"

#Development
salmon="ffa07a"
lightgreen="006400"#"e3fbe3"
failurejobslist=list()
fail_servertime=list()
fail_rundate=list()
fail_jobstatus=list()
def sheetchange(rowcount,tosheetname,fromsheetname,dictname):
    for i in range(2, rowcount+1):
        if fromsheetname.cell(row=i, column=4).value in ["Store Procedure [SP_MailManagementLegalHoldPopulate] to run everyday at 5 AM CST"]:
            tosheetname.cell(row=i, column=1).value = "LAPS-MailManagementLegalHoldPopulate"  # jobname
        else:
            tosheetname.cell(row=i, column=1).value = fromsheetname.cell(row=i, column=4).value
        #tosheetname.cell(row=i, column=1).value = fromsheetname.cell(row=i, column=4).value  # jobname
        tosheetname.cell(row=i, column=2).value = dictname[tosheetname.cell(row=i, column=1).value]  # servertime
        rundate = str(fromsheetname.cell(row=i, column=9).value)
        newrundate = rundate[6:8] + "/" + rundate[4:6] + "/" + rundate[0:4]
        tosheetname.cell(row=i, column=3).value = newrundate  # rundate
        if fromsheetname.cell(row=i , column=8).value == 1:
            tosheetname.cell(row=i, column=4).value = "success"  # jobstatus
            tosheetname["D" + str(i )].fill = PatternFill(start_color=lightgreen, end_color=lightgreen, fill_type="lightUp")
        elif fromsheetname.cell(row=i, column=8).value == 0:
            failurejobslist.append(fromsheetname.cell(row=i, column=4).value)
            fail_servertime.append(dictname[tosheetname.cell(row=i, column=1).value])
            rundate = str(fromsheetname.cell(row=i, column=9).value)
            newrundate = rundate[6:8] + "/" + rundate[4:6] + "/" + rundate[0:4]
            tosheetname.cell(row=i, column=3).value = newrundate
            fail_rundate.append(newrundate)
            fail_jobstatus.append("failure")

            tosheetname.cell(row=i, column=4).value = "failure"
            tosheetname["D" + str(i)].fill = PatternFill(start_color=salmon, end_color=salmon, fill_type="lightUp")

sheetchange(mc,lai,sh,laia_servertime)
sheetchange(mp,lap,sp,laps_servertime)
sheetchange(mz,lacc,sc,lacc_servertime)
#Headings
def head(coordinate,title):
    lai[coordinate]=title
    lap[coordinate] = title
    lacc[coordinate] = title
head("A1","Job Name")
head("B1","Server Time(CST)")
head("C1","Last run date")
head("D1","Job Status")

#setting column and row width and hight respectively
alphabets_string=string.ascii_uppercase
def row_col_dimension(dict_name,sheetname,rowcount):
    #for i in range(0,mr+1):
        #sheetname.column_dimensions[alphabets_string[i]].width=50#len(sheetname.cell(row=4,column=i+1).value)+10
    for i in range(1,rowcount+1):
        sheetname.row_dimensions[i].height=22
    sheetname.column_dimensions["A"].width = 70
    sheetname.column_dimensions["B"].width = 30
    sheetname.column_dimensions["C"].width = 15
    sheetname.column_dimensions["D"].width = 15
row_col_dimension(laia_shift,lai,mc)
row_col_dimension(laps_shift,lap,mp)
row_col_dimension(lacc_shift,lacc,mz)
#lai.column_dimensions["A"].width=70
#lai.column_dimensions["B"].width=50
#lai.column_dimensions["C"].width=18
#lai.column_dimensions["D"].width=18

#Setting Alignment of the headings in the cell and providing colur to cell  which contain headings
def align_ment_cellfill(coordinate_name):
    thistle = "D8BFD8"
    light_skyblue = "87CEFA"
    #Alignment
    #Aligning values to centre from horizontal and vertical side of the cell after merging cells
    #in this case copy has to be imported "from copy import copy"
    #for laia
    ali=copy(lai[coordinate_name].alignment)
    ali.horizontal='center'
    ali.vertical='center'
    lai[coordinate_name].alignment=ali
    #for laps
    ali1 = copy(lap[coordinate_name].alignment)
    ali1.horizontal = 'center'
    ali1.vertical = 'center'
    lap[coordinate_name].alignment = ali1
    # for lacc
    ali2 = copy(lacc[coordinate_name].alignment)
    ali2.horizontal = 'center'
    ali2.vertical = 'center'
    lacc[coordinate_name].alignment = ali2

    #lai[coordinate_name].alignment = Alignment(horizontal='center')
    #lap[coordinate_name].alignment = Alignment(horizontal='center')
    #Cell colur fill
    lai[coordinate_name].fill = PatternFill(start_color=thistle, end_color=thistle, fill_type="lightUp")
    lap[coordinate_name].fill = PatternFill(start_color=thistle, end_color=thistle, fill_type="lightUp")
    lacc[coordinate_name].fill = PatternFill(start_color=thistle, end_color=thistle, fill_type="lightUp")
align_ment_cellfill("A1")
align_ment_cellfill("B1")
align_ment_cellfill("C1")
align_ment_cellfill("D1")
#setting cell borders
mlai=lai.max_row
mlap=lap.max_row
mlacc=lacc.max_row
black = "000000"
thin = Side(border_style="thin", color=black)
def borderstyle(dictname,rowlength,sheetname,rowcount):
    #for j in range(0,mr+1):
    for j in range(0,4):
        for i in range(1,rowcount+1):
            sheetname[alphabets_string[j] + str(i)].border = Border(top=thin, left=thin, right=thin,bottom=thin)
borderstyle(laia_shift,mlai,lai,mc)
borderstyle(laps_shift,mlap,lap,mp)
borderstyle(lacc_shift,mlacc,lacc,mz)
w.save(path+"legal.xlsx")
#print("success")
joined_jobs="\n".join(failurejobslist)
#result = {}

#result["status"] = "success"
#result["failed_jobs"] = joined_jobs

failworkbook=openpyxl.Workbook()
failsheet=failworkbook.active
def fail(listname,colnum):
    for i in range(2,len(failurejobslist)+2):
        failsheet.cell(row=i,column=colnum).value=listname[i-2]
fail(failurejobslist,1)
fail(fail_servertime,2)
fail(fail_rundate,3)
fail(fail_jobstatus,4)
failsheet["A1"]="Job Name"
failsheet["B1"]="Server Time(CST)"
failsheet["C1"]="Last Run Date"
failsheet["D1"]="Job Status"

failworkbook.save(path+"failure.xlsx")


#output = {'output':result,'additional_attributes':result}
#sys.stdout.write(str(output)+'\n')
init(autoreset=True)
if len(joined_jobs)==0:
    print(Fore.RED+"All jobs are executed successfully")
else:
    print(joined_jobs)


#python laia.py C:\Users\2040664\anuraj\LEGAL\laiaquerydump.xlsx C:\Users\2040664\anuraj\LEGAL\lapsquerydump.xlsx C:\Users\2040664\anuraj\LEGAL\
